import webbrowser
import bs4
import requests
import openpyxl as xl
from myScrapingLib import getSoup

def headCellName(headCell):
	colRep = headCell.find('a')
	if(colRep is None):
		return headCell.text.strip()
	else:
		return colRep.attrs['title']

def entryCellVal(entryCell):
	cellLink = entryCell.find('a')
	if(cellLink is None):
		return entryCell.text
	else:
		return cellLink.attrs['title']

wikiNS = 'http://oldschoolrunescape.wikia.com'
tableIndex = '/wiki/Category:Slot_tables'

tableIndexSoup = getSoup(wikiNS+tableIndex) #bs4.BeautifulSoup(tableIndexPage.content, features='html5lib')

titledLinks = tableIndexSoup.select('a[title]')
slotTableLinks = [link for link in titledLinks if link.attrs['title'].endswith('slot table')]

tableDict = {}
for link in slotTableLinks:
	attrs = link.attrs	
	title = attrs['title']
	href = attrs['href']
	slotType = title.split(' ')[0]
	slotTableLink =  wikiNS + href
	tableDict[slotType] = slotTableLink
	#print('Slot: ' + slotType)
	#print('full link: ' + slotTableLink)
	
# Made into dict to remove dupes
print('From		dict of table page uris')
print('Making	dict of table heads and body elems')
print('Making	excel db')

WORKBOOK_NAME = 'OSRS_ARMORY'
wb = xl.Workbook()
# remove first sheet
wb.remove(wb.worksheets[0])

for (key,value) in tableDict.items():
	print(key + ' : ' + value)
	tablePage = getSoup(value)
	tableInPage = tablePage.select('table[class~=wikitable]')[0]
	body = tableInPage.find('tbody')
	
	# assume header is first row in tbody with no 'td'
	head = body.find(lambda tag: len(tag.find_all('td')) == 0)
	headColumns = head.find_all('th')
	num_columns = len(headColumns)

	# assume entries are rows in tbody with all 'td'
	# size of each entry must match size of header
	tableEntries = body.find_all(lambda tag: len(tag.find_all('td'))==num_columns)
	print(len(tableEntries))

	# Setup the excel sheet
	SHEET_NAME = key
	wb.create_sheet(SHEET_NAME)
	ws = wb[SHEET_NAME]

	ws['A1'] = value # store the uri to the page here
	for row in ws.iter_rows(min_row=1, min_col=2, max_col=2+num_columns-1, max_row=1):
		for cell in row:
			cell.value = headCellName(headColumns.pop(0))
	
	for (index, entry) in enumerate(tableEntries):
		print(index)
		rowNum = index+2
		entryCells = entry.find_all('td')
		#assume first entry cell has link out
		ws.cell(row=rowNum, column=1).value = wikiNS + entryCells[0].find('a').attrs['href']
		for colNum in range(2,num_columns+2):
			print(colNum)
			currCell = entryCells.pop(0)
			cellVal = entryCellVal(currCell)
			#print(cellVal)
			ws.cell(row=rowNum, column=colNum).value = cellVal



wb.save(WORKBOOK_NAME+'.xlsx')